import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

def generate_excel(output_file):
    sub_headers = [
        "STT", "ID", "Họ và tên", "Giới tính", "Ngày sinh",
        "Quản lý cấp 1", "Quản lý cấp 2", "Vị trí", "Loại HĐ", "HĐ Hiệu lực", "HĐ Hết hạn", "Hạn đánh giá",
        "Ngày bắt đầu", "Ngày nghỉ",
        "Thâm niên",
        "Địa chỉ", "Điện thoại", "Email cá nhân", "Số CCCD", "Ngày cấp", "Nơi cấp",
        "Họ tên (NT)", "Quan hệ", "Điện thoại (NT)",
        "Ngôn ngữ", "Chuyên ngành", "Trường ĐT", "Hệ", "Năm tốt nghiệp", "Chứng chỉ khác"
    ]

    # Generating 5 fresh employees with unique names to avoid duplicates
    data = [
        [1, "", "Trịnh Công Sơn", "Nam", "28/02/1991", "Manager Name", "", "Chuyên gia bảo mật", "Toàn thời gian", "01/01/2026", "01/01/2029", "", "01/01/2026", "", "Mới", "12 Trần Hưng Đạo, Quận 5, Tp. HCM", "0921000111", "son.trinh@security.com", "079091005555", "10/05/2020", "CA Tp. HCM", "Lê Thu Hà", "Vợ", "0921999888", "English, Java", "An toàn thông tin", "ĐH CNTT", "Chính quy", 2013, "CISSP, CEH"],
        [2, "", "Nguyễn Thị Ánh Viên", "Nữ", "15/11/1999", "Đỗ Thùy Linh", "", "Nhân viên Content", "Toàn thời gian", "01/04/2026", "01/04/2027", "", "01/04/2026", "", "Mới", "89 Lê Văn Sỹ, Phú Nhuận, Tp. HCM", "0931222333", "vien.nguyen@marketing.vn", "079200001234", "15/02/2022", "CA Tp. HCM", "Nguyễn Văn Nam", "Bố", "0931444555", "English", "Truyền thông", "ĐH KHXH&NV", "Chính quy", 2021, "Google Ads Certified"],
        [3, "", "Lương Xuân Trường", "Nam", "20/04/1995", "Phạm Minh Hoàng", "", "Dev Mobile", "Toàn thời gian", "01/05/2026", "01/05/2029", "", "01/05/2026", "", "Mới", "22 Tuyên Quang, Tp. Tuyên Quang", "0988776655", "truong.luong@mobile.dev", "001095012345", "12/03/2014", "CA Tuyên Quang", "Lương Văn Hùng", "Bố", "0988112233", "Korean, English", "Kỹ thuật phần mềm", "ĐH FPT", "Chính quy", 2017, "AWS Cloud Practitioner"],
        [4, "", "Hồ Ngọc Hà", "Nữ", "25/11/1984", "Director", "", "Manager Marketing", "Không thời hạn", "01/01/2023", "", "", "01/01/2023", "", "3 năm", "78 Lê Lợi, Quận 1, Tp. HCM", "0909000888", "ha.ho@creative.com", "001084001122", "15/05/2012", "CA Hà Nội", "Cường Đô La", "Chồng", "0909111222", "English", "Quản trị kinh doanh", "ĐH Ngoại Thương", "Chính quy", 2006, "MBA Excellence"],
        [5, "", "Trần Thành", "Nam", "05/02/1987", "Director", "", "Kỹ sư hệ thống", "Toàn thời gian", "15/05/2026", "15/05/2029", "", "15/05/2026", "", "Mới", "102 Nguyễn Trãi, Quận 1, Tp. HCM", "0944556677", "thanh.tran@systems.com", "036087009988", "20/08/2015", "CA Nam Định", "Hari Won", "Vợ", "0944112233", "English, Chinese", "Điện tử viễn thông", "ĐH Bách Khoa HN", "Chính quy", 2009, "PMP, Azure Solutions Architect"]
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách nhân viên"

    # Define Styles
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header creation (same logic as before to match system template)
    ws.append(sub_headers) 
    for i in range(1, 31):
        ws.cell(row=1, column=i).border = thin_border
        ws.cell(row=1, column=i).alignment = center_align
        ws.cell(row=1, column=i).font = bold_font

    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=14)
    ws.cell(row=1, column=6).value = "THÔNG TIN HỢP ĐỒNG HIỆN TẠI"
    ws.merge_cells(start_row=1, start_column=16, end_row=1, end_column=21)
    ws.cell(row=1, column=16).value = "THÔNG TIN CÁ NHÂN"
    ws.merge_cells(start_row=1, start_column=22, end_row=1, end_column=24)
    ws.cell(row=1, column=22).value = "THÔNG TIN NGƯỜI THÂN"
    ws.merge_cells(start_row=1, start_column=25, end_row=1, end_column=30)
    ws.cell(row=1, column=25).value = "TRÌNH ĐỘ"

    ws.insert_rows(2)
    for i, head in enumerate(sub_headers):
        ws.cell(row=2, column=i+1).value = head
        ws.cell(row=2, column=i+1).border = thin_border
        ws.cell(row=2, column=i+1).alignment = center_align
        ws.cell(row=2, column=i+1).font = bold_font

    ws.insert_rows(3)
    for i in range(1, 31):
        if i == 5: ws.cell(row=3, column=i).value = "[dd/mm/yyyy]"
        else: ws.cell(row=3, column=i).value = i
        ws.cell(row=3, column=i).border = thin_border
        ws.cell(row=3, column=i).alignment = center_align

    # Add 5 NEW Data rows
    for row_data in data:
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Adjust column widths
    for i in range(1, 31):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = 22

    wb.save(output_file)
    print(f"File '{output_file}' has been updated with 5 NEW employees.")

if __name__ == "__main__":
    generate_excel("fake_5_employees_v3.xlsx")
