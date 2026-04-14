import requests
import json
import base64
import os

# Install openpyxl if needed: os.system("pip install openpyxl")
try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

BASE_URL = "http://localhost:8080/api"

def main():
    print("1. Logging in as Admin...")
    login_url = f"{BASE_URL}/auth/login"
    
    session = requests.Session()
    res = session.post(login_url, json={
        "email": "admin@company.com",
        "password": "Admin@123"
    })
    
    if res.status_code != 200:
        print("Login failed:", res.text)
        return
        
    print("Login successful. Cookies set:", session.cookies.get_dict())

    print("2. Downloading Excel Template...")
    template_url = f"{BASE_URL}/employees/template"
    res = session.get(template_url)
    if res.status_code != 200:
        print("Failed to download template (Status):", res.status_code)
        print("Reason:", res.reason)
        return
        
    template_path = "template.xlsx"
    with open(template_path, "wb") as f:
        f.write(res.content)
    print(f"Template downloaded to {template_path}")

    print("3. Modifying Template to add 3 fake employees...")
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active

    # Template already has Row 4 (index 3 in java, index 4 in openpyxl) as "Nguyễn Văn Tuấn"
    # Row 4 is row with values.
    # Let's read Row 4's values
    base_row = [sheet.cell(row=4, column=c).value for c in range(1, 31)]

    # Employee 1: (From template, just ensure uniqueness)
    base_row[0] = 1 # STT
    base_row[2] = "Nguyễn Hữu Tài" # Name
    base_row[4] = "12/05/2000"
    base_row[16] = "0911223344" # Phone
    base_row[18] = "079011223344" # CCCD
    for c in range(1, 31):
        sheet.cell(row=4, column=c, value=base_row[c-1])

    # Employee 2:
    row2 = base_row.copy()
    row2[0] = 2
    row2[2] = "Đặng Hoàng Yến"
    row2[3] = "Nữ"
    row2[16] = "0922334455"
    row2[18] = "079022334455"
    for c in range(1, 31):
        sheet.cell(row=5, column=c, value=row2[c-1])

    # Employee 3:
    row3 = base_row.copy()
    row3[0] = 3
    row3[2] = "Vũ Tuấn Kiệt"
    row3[16] = "0933445566"
    row3[18] = "079033445566"
    for c in range(1, 31):
        sheet.cell(row=6, column=c, value=row3[c-1])

    fake_file_path = "fake_3_employees_v2.xlsx"
    wb.save(fake_file_path)
    print(f"Fake file saved to {fake_file_path} ! Ready for you to upload via UI.")

if __name__ == "__main__":
    main()
