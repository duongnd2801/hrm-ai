import openpyxl
import psycopg2
import random
from datetime import datetime, timedelta

def get_workdays(start_date, end_date):
    workdays = []
    curr = start_date
    while curr <= end_date:
        if curr.weekday() < 5: # Monday to Friday
            workdays.append(curr.strftime('%d/%m/%Y'))
        curr += timedelta(days=1)
    return workdays

try:
    # 1. Connect to DB to get employees
    conn = psycopg2.connect(
        dbname='hrm_db',
        user='postgres',
        password='1234$', # Fixed from 1234$ to match script
        host='localhost', port='5432'
    )
    conn.set_client_encoding('UTF8')
    cur = conn.cursor()
    # Get up to 30 employees to generate mass data
    cur.execute("SELECT CAST(id AS TEXT), full_name FROM hrm.employees LIMIT 30;")
    employees = cur.fetchall()
    cur.close()
    conn.close()

    if not employees:
        print("No employees found in database. Please import employees first.")
        exit()

    # 2. Load or Create Excel
    file_path = 'cham_cong_nhan_vien.xlsx'
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["STT", "ID Nhân viên", "Họ tên", "Phòng ban", "Ngày", "Máy", "Check-in", "Check-out"])

    # Find the last row and last STT
    last_row = ws.max_row
    last_stt = 0
    if last_row > 1:
        try:
             last_stt = int(ws.cell(row=last_row, column=1).value)
        except:
             last_stt = last_row - 1

    # 3. Define Dates (May 04 to May 15, 2026)
    start_dt = datetime(2026, 5, 4)
    end_dt = datetime(2026, 5, 15)
    new_dates = get_workdays(start_dt, end_dt)
    
    current_row = last_row + 1
    stt_counter = last_stt + 1
    
    records_added = 0
    for emp_id, full_name in employees:
        for date_str in new_dates:
            # Check-in: 08:00 - 09:15
            ci_hour = random.randint(8, 8)
            ci_min = random.randint(15, 59)
            if random.random() > 0.8: # 20% chance of being slightly late
                ci_hour = 9
                ci_min = random.randint(0, 15)
            
            # Check-out: 17:30 - 19:30
            co_hour = random.randint(17, 18)
            co_min = random.randint(30, 59)
            if random.random() > 0.7: # 30% chance of OT
                co_hour = 19
                co_min = random.randint(0, 45)

            ws.cell(row=current_row, column=1).value = stt_counter
            ws.cell(row=current_row, column=2).value = emp_id
            ws.cell(row=current_row, column=3).value = full_name
            ws.cell(row=current_row, column=4).value = "Công ty HRM"
            ws.cell(row=current_row, column=5).value = date_str
            ws.cell(row=current_row, column=6).value = "Máy 01"
            ws.cell(row=current_row, column=7).value = f"{ci_hour:02d}:{ci_min:02d}"
            ws.cell(row=current_row, column=8).value = f"{co_hour:02d}:{co_min:02d}"
            
            current_row += 1
            stt_counter += 1
            records_added += 1

    wb.save(file_path)
    print(f"Success! Added {records_added} records for {len(employees)} employees across {len(new_dates)} workdays.")
    print(f"Dates covered: {new_dates[0]} to {new_dates[-1]}")

except Exception as e:
    print(f"Error: {e}")
